"""
Excel Exporter for generating formatted Excel spreadsheets
"""

import logging
from typing import List, Optional, Any
from pathlib import Path
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from PIL import Image as PILImage

from models.general import General
from utils.resource_manager import resource_manager

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Handles Excel export functionality"""

    def __init__(self, config: dict):
        self.config = config
        self.template_path = config.get('excel_template_path', 'Resources/EvonyActiveGenerals.xltx')

    def export_generals(self, generals: List[General], file_path: str, count_text: str = "", incremental: bool = False) -> bool:
        """Export generals data to Excel file"""
        try:
            logger.info(f"Exporting {len(generals)} generals to {file_path} (incremental={incremental})")
            
            # Debug: Log general data
            for i, gen in enumerate(generals):
                logger.debug(f"General {i+1}: name='{gen.name}', level={gen.level}, power={gen.power}, has_cultivation={bool(gen.cultivation_data)}, has_specialty={bool(gen.specialty_names)}, has_covenant={bool(gen.covenant_data)}")

            # For incremental exports, load existing file if it exists
            if incremental and Path(file_path).exists():
                logger.info(f"Loading existing incremental file: {file_path}")
                workbook = openpyxl.load_workbook(file_path)
            else:
                # Create or load workbook from template
                workbook = self.create_workbook(self.template_path)
                
            if not workbook:
                logger.error("Failed to create/load workbook")
                return False

            worksheet = workbook.active
            if not worksheet:
                worksheet = workbook.create_sheet("Generals")

            # Write count text to B3 if provided (only for non-incremental)
            if count_text and not incremental:
                worksheet.cell(row=3, column=2, value=count_text)
                logger.info(f"Wrote count text to B3: {count_text}")

            # Clear existing data rows (keep headers) - only for non-incremental
            if not incremental:
                self.clear_data_rows(worksheet)

            # Populate with general data
            self.populate_data(worksheet, generals)

            # Apply formatting
            self.format_cells(worksheet, len(generals))

            # Insert images
            self.insert_images(worksheet, generals)

            # Save file
            workbook.save(file_path)
            logger.info(f"Excel export completed: {file_path}")

            return True

        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return False

    def append_general(self, general: General, file_path: str, row_index: int) -> bool:
        """Append a single general to an existing Excel file"""
        try:
            logger.info(f"Appending general '{general.name}' to row {row_index} in {file_path}")
            
            # Load existing workbook
            if not Path(file_path).exists():
                logger.error(f"Excel file does not exist: {file_path}")
                return False
                
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            if not worksheet:
                logger.error("No active worksheet found")
                return False

            # Populate single general data
            self.populate_single_general(worksheet, general, row_index)

            # Insert single general images
            self.insert_single_general_images(worksheet, general, row_index)

            # Set row height to 1 inch (72 points) with content adjustment
            max_height = 72  # Default height: 1 inch (72 points)
            for col in range(1, 16):
                cell_value = worksheet.cell(row=row_index, column=col).value
                if cell_value:
                    # Estimate height based on text length (rough approximation)
                    text_length = len(str(cell_value))
                    # For multi-line content (like specialty data), estimate lines
                    if '\n' in str(cell_value):
                        lines = str(cell_value).count('\n') + 1
                        estimated_height = lines * 15  # 15 points per line
                    else:
                        # Single line, but wrap if very long
                        estimated_height = min(text_length // 50 + 1, 5) * 15
                    max_height = max(max_height, estimated_height)
            worksheet.row_dimensions[row_index].height = max_height

            # Save file
            workbook.save(file_path)
            logger.info(f"Successfully appended general to {file_path}")

            return True

        except Exception as e:
            logger.error(f"Failed to append general: {e}")
            return False

    def populate_single_general(self, worksheet: Any, general: General, row: int) -> None:
        """Populate a single general's data in worksheet"""
        try:
            logger.debug(f"Populating row {row} with general: name={general.name}, level={general.level}, power={general.power}")
            logger.debug(f"  exp_ratio={general.exp_ratio}, cultivation={general.cultivation_data[:50] if general.cultivation_data else 'None'}")
            logger.debug(f"  specialty={general.specialty_names[:50] if general.specialty_names else 'None'}")
            logger.debug(f"  covenant={general.covenant_data[:50] if general.covenant_data else 'None'}")
            
            worksheet.cell(row=row, column=1, value=general.name or "")
            worksheet.cell(row=row, column=2, value=general.level)
            worksheet.cell(row=row, column=3, value=general.type or "")
            worksheet.cell(row=row, column=4, value=general.power)
            # Stars column (5) will have image
            worksheet.cell(row=row, column=6, value=general.exp_ratio or "")
            
            # Split cultivation data into separate columns
            if general.cultivation_data:
                cultivation_parts = general.cultivation_data.split('\n')
                worksheet.cell(row=row, column=7, value=cultivation_parts[0] if len(cultivation_parts) > 0 else "")  # Leadership
                worksheet.cell(row=row, column=8, value=cultivation_parts[1] if len(cultivation_parts) > 1 else "")  # Attack
                worksheet.cell(row=row, column=9, value=cultivation_parts[2] if len(cultivation_parts) > 2 else "")  # Defense
                worksheet.cell(row=row, column=10, value=cultivation_parts[3] if len(cultivation_parts) > 3 else "") # Politics
            else:
                worksheet.cell(row=row, column=7, value="")
                worksheet.cell(row=row, column=8, value="")
                worksheet.cell(row=row, column=9, value="")
                worksheet.cell(row=row, column=10, value="")
            
            # Specialties - use combined names
            specialty_cell = worksheet.cell(row=row, column=11)
            if general.specialty_names:
                specialty_cell.value = general.specialty_names
                specialty_cell.number_format = '@'  # Text format
                specialty_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                specialty_cell.value = ""
            
            # Covenant - use combined names
            covenant_cell = worksheet.cell(row=row, column=13)
            if general.covenant_names:
                covenant_cell.value = general.covenant_names
                covenant_cell.number_format = '@'  # Text format
                covenant_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                covenant_cell.value = ""
            
            worksheet.cell(row=row, column=15, value="Yes" if general.is_uncertain else "No")

            # Apply alignments to other cells
            # Top vertical for A,B,D,F,G,H,I,J,K,N (1,2,4,6,7,8,9,10,11,14)
            # Center horizontal for B,D,F,G,H,I,J (2,4,6,7,8,9,10)
            for col in [1, 2, 4, 6, 7, 8, 9, 10, 11, 14]:
                cell = worksheet.cell(row=row, column=col)
                if col in [2, 4, 6, 7, 8, 9, 10]:
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top")

            logger.debug(f"Populated single general data in row {row}")
        except Exception as e:
            logger.error(f"Failed to populate single general data: {e}")

    def insert_single_general_images(self, worksheet: Any, general: General, row: int) -> None:
        """Insert images for a single general"""
        try:
            # Add stars image
            if general.stars_image and len(general.stars_image) > 0:
                try:
                    # Create image from bytes
                    image_stream = BytesIO(general.stars_image)
                    img = XLImage(image_stream)

                    # Add image to worksheet
                    worksheet.add_image(img, f'E{row}')

                except Exception as img_error:
                    logger.warning(f"Failed to insert stars image for general {general.name}: {img_error}")

            # Add type image
            if general.type_image and len(general.type_image) > 0:
                try:
                    # Create image from bytes
                    image_stream = BytesIO(general.type_image)
                    img = XLImage(image_stream)

                    # Add image to worksheet, centered in cell
                    worksheet.add_image(img, f'C{row}')

                except Exception as img_error:
                    logger.warning(f"Failed to insert type image for general {general.name}: {img_error}")

            # Add specialty combined image in column L
            if general.specialty_combined_image and len(general.specialty_combined_image) > 0:
                try:
                    # Create image from bytes
                    image_stream = BytesIO(general.specialty_combined_image)
                    img = XLImage(image_stream)

                    # Add image to worksheet, centered in cell
                    worksheet.add_image(img, f'L{row}')

                except Exception as img_error:
                    logger.warning(f"Failed to insert specialty combined image for general {general.name}: {img_error}")

            # Add covenant combined image in column N
            if general.covenant_combined_image and len(general.covenant_combined_image) > 0:
                try:
                    # Create image from bytes
                    image_stream = BytesIO(general.covenant_combined_image)
                    img = XLImage(image_stream)

                    # Add image to worksheet, centered in cell
                    worksheet.add_image(img, f'N{row}')

                except Exception as img_error:
                    logger.warning(f"Failed to insert covenant combined image for general {general.name}: {img_error}")

            # Add covenant attributes image in column O
            if general.covenant_attributes_image and len(general.covenant_attributes_image) > 0:
                try:
                    # Create image from bytes
                    image_stream = BytesIO(general.covenant_attributes_image)
                    img = XLImage(image_stream)

                    # Add image to worksheet, centered in cell
                    worksheet.add_image(img, f'O{row}')

                except Exception as img_error:
                    logger.warning(f"Failed to insert covenant attributes image for general {general.name}: {img_error}")

            logger.debug(f"Inserted images for single general in row {row}")
        except Exception as e:
            logger.error(f"Failed to insert single general images: {e}")

    def create_workbook(self, template_path: Optional[str] = None) -> Optional[Any]:
        """Create or load Excel workbook"""
        try:
            # Try to load template using resource manager
            if template_path:
                resource_path = resource_manager.get_resource_path(template_path)
                if resource_path.exists():
                    logger.info(f"Loading template from resources: {resource_path}")
                    return openpyxl.load_workbook(str(resource_path))
                else:
                    logger.warning(f"Template not found in resources: {template_path}")

            # Fallback: try direct path for development
            if template_path and Path(template_path).exists():
                logger.info(f"Loading template from direct path: {template_path}")
                return openpyxl.load_workbook(template_path)
            else:
                logger.info("Creating new workbook")
                workbook = openpyxl.Workbook()
                # Create headers
                worksheet = workbook.active
                headers = [
                    "Name", "Level", "Type", "Power", "Stars", "Experience Ratio",
                    "Leadership", "Attack", "Defense", "Politics", "Specialties", "Specialty Images",
                    "Covenants", "Covenant Images", "Covenant Attributes", "Uncertain"
                ]
                for col, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=col, value=header)
                return workbook
        except Exception as e:
            logger.error(f"Failed to create workbook: {e}")
            return None

    def combine_images_side_by_side(self, image_bytes_list: List[bytes]) -> bytes:
        """Combine multiple images side by side into a single image"""
        if not image_bytes_list:
            return b""
        
        if len(image_bytes_list) == 1:
            return image_bytes_list[0]
        
        try:
            # Open all images
            images = []
            for img_bytes in image_bytes_list:
                if img_bytes:
                    img_stream = BytesIO(img_bytes)
                    img = PILImage.open(img_stream)
                    images.append(img)
            
            if not images:
                return b""
            
            # Find the maximum height
            max_height = max(img.height for img in images)
            
            # Resize all images to the same height while maintaining aspect ratio
            resized_images = []
            for img in images:
                if img.height != max_height:
                    # Calculate new width to maintain aspect ratio
                    aspect_ratio = img.width / img.height
                    new_width = int(max_height * aspect_ratio)
                    resized_img = img.resize((new_width, max_height), PILImage.Resampling.LANCZOS)
                    resized_images.append(resized_img)
                else:
                    resized_images.append(img)
            
            # Calculate total width
            total_width = sum(img.width for img in resized_images)
            
            # Create new image with combined width
            combined_image = PILImage.new('RGBA', (total_width, max_height))
            
            # Paste images side by side
            current_x = 0
            for img in resized_images:
                combined_image.paste(img, (current_x, 0))
                current_x += img.width
            
            # Convert back to bytes
            output_stream = BytesIO()
            combined_image.save(output_stream, format='PNG')
            return output_stream.getvalue()
            
        except Exception as e:
            logger.error(f"Failed to combine images: {e}")
            return b""

    def load_covenant_attributes_image(self) -> bytes:
        """Load the GeneralsListCovenantAttributes.png image"""
        try:
            # Try resource manager first
            attributes_path = "Resources/GeneralsListCovenantAttributes.png"
            image_data = resource_manager.read_resource_bytes(attributes_path)
            if image_data:
                return image_data

            # Fallback: try direct path for development
            direct_path = Path(self.config.get('resources_path', 'Resources')) / "GeneralsListCovenantAttributes.png"
            if direct_path.exists():
                with open(direct_path, 'rb') as f:
                    return f.read()
            else:
                logger.warning(f"Covenant attributes image not found: {direct_path}")
                return b""
        except Exception as e:
            logger.error(f"Failed to load covenant attributes image: {e}")
            return b""

    def clear_data_rows(self, worksheet: Any) -> None:
        """Clear existing data rows in workbook"""
        try:
            # Clear from row 7 onwards (keep headers in rows 1-6)
            for row in range(7, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=row, column=col).value = None
            logger.info(f"Cleared data rows from 7 to {worksheet.max_row}")
        except Exception as e:
            logger.error(f"Failed to clear data rows: {e}")

    def populate_data(self, worksheet: Any, generals: List[General]) -> None:
        """Populate worksheet with general data"""
        try:
            for row, general in enumerate(generals, 7):  # Start from row 7
                logger.debug(f"Populating row {row} with general: name={general.name}, level={general.level}, power={general.power}")
                logger.debug(f"  exp_ratio={general.exp_ratio}, cultivation={general.cultivation_data[:50] if general.cultivation_data else 'None'}")
                logger.debug(f"  specialty={general.specialty_names[:50] if general.specialty_names else 'None'}")
                logger.debug(f"  covenant={general.covenant_data[:50] if general.covenant_data else 'None'}")
                
                worksheet.cell(row=row, column=1, value=general.name or "")
                worksheet.cell(row=row, column=2, value=general.level)
                worksheet.cell(row=row, column=3, value=general.type or "")
                worksheet.cell(row=row, column=4, value=general.power)
                # Stars column (5) will have image
                worksheet.cell(row=row, column=6, value=general.exp_ratio or "")
                
                # Split cultivation data into separate columns
                if general.cultivation_data:
                    cultivation_parts = general.cultivation_data.split('\n')
                    worksheet.cell(row=row, column=7, value=cultivation_parts[0] if len(cultivation_parts) > 0 else "")  # Leadership
                    worksheet.cell(row=row, column=8, value=cultivation_parts[1] if len(cultivation_parts) > 1 else "")  # Attack
                    worksheet.cell(row=row, column=9, value=cultivation_parts[2] if len(cultivation_parts) > 2 else "")  # Defense
                    worksheet.cell(row=row, column=10, value=cultivation_parts[3] if len(cultivation_parts) > 3 else "") # Politics
                else:
                    worksheet.cell(row=row, column=7, value="")
                    worksheet.cell(row=row, column=8, value="")
                    worksheet.cell(row=row, column=9, value="")
                    worksheet.cell(row=row, column=10, value="")
                
                # Specialties - use combined names
                if general.specialty_names:
                    worksheet.cell(row=row, column=11, value=general.specialty_names)
                else:
                    worksheet.cell(row=row, column=11, value="")
                
                # Covenant - use combined names
                if general.covenant_names:
                    worksheet.cell(row=row, column=13, value=general.covenant_names)
                else:
                    worksheet.cell(row=row, column=13, value="")
                
            logger.info(f"Populated data for {len(generals)} generals")
        except Exception as e:
            logger.error(f"Failed to populate data: {e}")

    def format_cells(self, worksheet: Any, num_generals: int) -> None:
        """Apply formatting to worksheet cells"""
        try:
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Format headers (now 15 columns)
            for col in range(1, 16):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Format data cells (start from row 7, 15 columns)
            for row in range(7, num_generals + 7):
                for col in range(1, 16):
                    cell = worksheet.cell(row=row, column=col)
                    # Columns K (11) and M (13) have word wrap enabled
                    if col in [11, 13]:
                        # For K and M, wrap text
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    elif col in [2, 4, 6, 7, 8, 9, 10]:
                        # Center horizontally for B,D,F,G,H,I,J
                        cell.alignment = Alignment(horizontal="center", vertical="top")
                    else:
                        # Top alignment for A,B,D,F,G,H,I,J,K,N (1,2,4,6,7,8,9,10,11,14)
                        cell.alignment = Alignment(horizontal="left", vertical="top")
                    cell.border = border

            # Apply number formatting to power column (column 4)
            from openpyxl.styles import NamedStyle
            number_style = NamedStyle(name='number_with_commas', number_format='#,##0')
            for row in range(7, num_generals + 7):
                cell = worksheet.cell(row=row, column=4)
                cell.style = number_style

            # Auto-adjust column widths (15 columns)
            for col in range(1, 16):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in range(1, num_generals + 7):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Auto-adjust row heights based on content (data rows only, preserve template header heights)
            for row in range(7, num_generals + 7):
                max_height = 72  # Default height: 1 inch (72 points)
                for col in range(1, 16):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        # Estimate height based on text length (rough approximation)
                        text_length = len(str(cell_value))
                        # For multi-line content (like specialty data), estimate lines
                        if '\n' in str(cell_value):
                            lines = str(cell_value).count('\n') + 1
                            estimated_height = lines * 15  # 15 points per line
                        else:
                            # Single line, but wrap if very long
                            estimated_height = min(text_length // 50 + 1, 5) * 15
                        max_height = max(max_height, estimated_height)
                worksheet.row_dimensions[row].height = max_height

            # Special widths for image columns
            worksheet.column_dimensions['C'].width = 15  # Type images
            worksheet.column_dimensions['E'].width = 20  # Stars images (fits ~140px wide images)
            worksheet.column_dimensions['L'].width = 20  # Specialty images
            worksheet.column_dimensions['N'].width = 20  # Covenant images
            worksheet.column_dimensions['O'].width = 20  # Covenant attributes image

            logger.info("Applied cell formatting")
        except Exception as e:
            logger.error(f"Failed to format cells: {e}")

    def insert_images(self, worksheet: Any, generals: List[General]) -> None:
        """Insert images into worksheet"""
        try:
            for row, general in enumerate(generals, 7):  # Start from row 7
                if general.stars_image and len(general.stars_image) > 0:
                    try:
                        logger.debug(f"Inserting stars image for {general.name}: {len(general.stars_image)} bytes")
                        # Create image from bytes
                        image_stream = BytesIO(general.stars_image)
                        img = XLImage(image_stream)

                        # Add image to worksheet
                        worksheet.add_image(img, f'E{row}')

                    except Exception as img_error:
                        logger.warning(f"Failed to insert stars image for general {general.name}: {img_error}")

                # Add type image if available
                if general.type_image and len(general.type_image) > 0:
                    try:
                        # Create image from bytes
                        image_stream = BytesIO(general.type_image)
                        img = XLImage(image_stream)

                        # Add image to worksheet, centered in cell
                        worksheet.add_image(img, f'C{row}')

                    except Exception as img_error:
                        logger.warning(f"Failed to insert type image for general {general.name}: {img_error}")

                # Add specialty combined image in column L
                if general.specialty_combined_image and len(general.specialty_combined_image) > 0:
                    try:
                        # Create image from bytes
                        image_stream = BytesIO(general.specialty_combined_image)
                        img = XLImage(image_stream)

                        # Add image to worksheet, centered in cell
                        worksheet.add_image(img, f'L{row}')

                    except Exception as img_error:
                        logger.warning(f"Failed to insert specialty combined image for general {general.name}: {img_error}")

                # Add covenant combined image in column N
                if general.covenant_combined_image and len(general.covenant_combined_image) > 0:
                    try:
                        # Create image from bytes
                        image_stream = BytesIO(general.covenant_combined_image)
                        img = XLImage(image_stream)

                        # Add image to worksheet, centered in cell
                        worksheet.add_image(img, f'N{row}')

                    except Exception as img_error:
                        logger.warning(f"Failed to insert covenant combined image for general {general.name}: {img_error}")

                # Add covenant attributes image in column O
                if general.covenant_attributes_image and len(general.covenant_attributes_image) > 0:
                    try:
                        # Create image from bytes
                        image_stream = BytesIO(general.covenant_attributes_image)
                        img = XLImage(image_stream)

                        # Add image to worksheet, centered in cell
                        worksheet.add_image(img, f'O{row}')

                    except Exception as img_error:
                        logger.warning(f"Failed to insert covenant attributes image for general {general.name}: {img_error}")

            logger.info(f"Inserted images for {len(generals)} generals")
        except Exception as e:
            logger.error(f"Failed to insert images: {e}")